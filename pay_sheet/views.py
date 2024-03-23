import datetime
import json
import locale
from pprint import pprint

from django.apps import apps
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Q
from django.http import Http404, JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView, ListView
from loguru import logger
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from completed_works.models import WorkRecord, WorkRecordQuantity, Standards
from pay_sheet.models import PaySheetModel, PaySheetMonthModel
from users.models import CustomUser, Role
from utils.utils import get_dates, get_days_for_current_and_next_month
from work_schedule.models import Appointment


class PaySheet(LoginRequiredMixin, TemplateView):
    template_name = 'pay_sheet/pay_sheet.html'
    login_url = '/users/login/'
    success_url = reverse_lazy('pay_sheet:pay_sheet')

    def create_excel_file(self, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Расчет'

        if self.template_name == 'pay_sheet/pay_sheet.html':
            pay_sheet_model = apps.get_model(app_label='pay_sheet', model_name='PaySheetModel')
        elif self.template_name == 'pay_sheet/pay_sheet_month.html':
            pay_sheet_model = apps.get_model(app_label='pay_sheet', model_name='PaySheetMonthModel')

        # Получите метаданные модели, чтобы получить поля
        model_fields = pay_sheet_model._meta.fields
        # Добавьте столбец "Role" к заголовкам столбцов
        headers = [field.verbose_name for field in model_fields]
        headers.append('Телефон')
        headers.append('Карта')

        # Создайте заголовки столбцов на основе полей модели
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Заполните данные из queryset
        for row_num, appointment in enumerate(queryset, 2):
            for col_num, field in enumerate(model_fields, 1):
                cell_value = getattr(appointment, field.name)

                # Получите должность пользователя и добавьте ее в конец строки
                if field.name == 'user' and isinstance(cell_value, CustomUser):
                    logger.debug(cell_value.phone_number)
                    worksheet.cell(row=row_num, column=col_num, value=cell_value.username)
                    worksheet.cell(row=row_num, column=len(headers) - 1, value=str(cell_value.phone_number))
                    worksheet.cell(row=row_num, column=len(headers), value=str(cell_value.card_details))
                elif isinstance(cell_value, datetime.datetime):
                    data_str = str(cell_value)
                    data_datetime = datetime.datetime.fromisoformat(data_str)
                    # Форматирование даты и времени в требуемый формат
                    formatted_str = data_datetime.strftime("%Y-%m-%d %H:%M:%S")
                    worksheet.cell(row=row_num, column=col_num, value=formatted_str)

                else:
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)

                max_length = len(str(cell_value))
                column_letter = get_column_letter(col_num)
                if worksheet.column_dimensions[column_letter].width is None or worksheet.column_dimensions[
                    column_letter].width < max_length:
                    worksheet.column_dimensions[column_letter].width = max_length

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=salary.xlsx'
        workbook.save(response)

        return response

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if 'download_excel' in request.GET:
            excel_response = self.create_excel_file(queryset)
            return excel_response

        context = self.get_context_data()
        return self.render_to_response(context)

    def get_queryset(self):
        if hasattr(self, '_queryset'):
            return self._queryset

        year = self.request.GET.get('year', None)
        week = self.request.GET.get('week', None)
        month = self.request.GET.get('month', None)

        if self.template_name == 'pay_sheet/pay_sheet.html':
            if not year or not week:
                import datetime
                today = datetime.date.today()
                year = today.year
                week = today.isocalendar()[1]
                if week != 1:
                    week -= 1
            queryset = PaySheetModel.objects.filter(
                year=year,
                week=week,
                user__role__type_salary=Role.TYPE_SALARY[1][0]
            )
            logger.debug(year)
            logger.debug(week)

        elif self.template_name == 'pay_sheet/pay_sheet_month.html':
            if not year or not month:
                import datetime
                today = datetime.date.today()
                year = today.year
                month = today.month
            queryset = PaySheetMonthModel.objects.filter(
                year=year,
                month=month,
                user__role__type_salary=Role.TYPE_SALARY[2][0]
            )
        self._queryset = queryset
        return queryset

    def get_context_data(self, **kwargs):
        logger.debug(self.template_name)
        if self.template_name == 'pay_sheet/pay_sheet.html':
            logger.success(Role.TYPE_SALARY[1][0])
            filter_role = Q(role__type_salary=Role.TYPE_SALARY[1][0])
            filter_role2 = Q(user__role__type_salary=Role.TYPE_SALARY[1][0])
        elif self.template_name == 'pay_sheet/pay_sheet_month.html':
            filter_role = Q(role__type_salary=Role.TYPE_SALARY[2][0])
            filter_role2 = Q(user__role__type_salary=Role.TYPE_SALARY[2][0])

        context = super().get_context_data(**kwargs)
        if not self.request.user.is_staff:
            raise Http404
        time_start = datetime.datetime.now()
        users_dict = {}
        total_salary = 0
        total_result_salary = 0

        logger.debug(self.request.GET)
        year = self.request.GET.get('year', None)
        week = self.request.GET.get('week', None)
        month = self.request.GET.get('month', None)
        today = datetime.date.today()
        current_year = today.year
        current_month = today.month
        current_week = today.isocalendar()[1]

        if not year or not week:
            year = today.year
            week = today.isocalendar()[1]
            if week != 1:
                week -= 1
        else:
            year, week = int(year), int(week)

        monday, sunday = get_dates(year, week)
        user_list = CustomUser.objects.filter(filter_role)
        logger.debug(user_list)
        for user in user_list:
            queryset_appointments = Appointment.objects.filter(user=user, date__year=year, date__week=week,
                                                               verified=True)
            if len(queryset_appointments) == 0:
                continue
            week_hours_work = []
            for i in range(7):
                date_day = monday + datetime.timedelta(days=i)
                try:
                    duration_day = sum([int(i.duration.total_seconds() / 3600) for i in
                                        queryset_appointments.filter(user=user, date=date_day)])
                    week_hours_work.append(duration_day)
                except Exception as ex:
                    week_hours_work.append(0)

            users_dict[user] = {'week_hours_work': week_hours_work}
            users_dict[user]['hours'] = sum(users_dict[user]['week_hours_work'])
            count_of_12 = users_dict[user]['week_hours_work'].count(12)
            users_dict[user]['count_of_12'] = count_of_12

            users_dict[user]['flag'] = True if count_of_12 >= 3 else False

            # Высчитывание зарплаты по часам
            salary = 0
            mess = ''

            if user.role.name == 'Упаковщик 2' or user.role.name == 'Упаковщик':
                if users_dict[user]['flag']:
                    for day_hour in users_dict[user]['week_hours_work']:
                        if day_hour == 12:
                            salary += 12 * 150
                        else:
                            salary += day_hour * 120
                else:
                    salary = sum([day_hour * 120 for day_hour in users_dict[user]['week_hours_work']])
                    mess += 'Не отработано 3 дня по 12ч.; '
            else:
                salary = sum([day_hour * user.role.salary for day_hour in users_dict[user]['week_hours_work']])

            users_dict[user]['salary'] = salary
            total_salary += salary
            # высчитывание коэффецента к зарплате
            try:
                if users_dict[user]['hours'] != 0:
                    work_totals_dict = calculate_work_totals_for_week(user, users_dict[user]['hours'], monday, sunday)
                    if len(work_totals_dict) != 0:
                        users_dict[user]['works'] = work_totals_dict
                        kf = sum([i[1] for i in work_totals_dict.values()]) * 100
                        users_dict[user]['kf'] = round(kf, 2)

                        if users_dict[user]['hours'] < 20:
                            mess += 'Отработанно меньше 20 часов; '
                        if kf >= 80:
                            result_salary = salary
                        else:
                            result_salary = salary * kf / 100
                            mess += 'Коэффецент меньше 80%; '

                        users_dict[user]['result_salary'] = round(result_salary, 2)
                    else:
                        users_dict[user]['works'] = {}
                        users_dict[user]['kf'] = 0
                        users_dict[user]['result_salary'] = 0
                        mess = 'Нет сдельных листов; '
                        if user.status_work == False:
                            mess += 'Уволен;'
                else:
                    users_dict[user]['works'] = {}
                    users_dict[user]['kf'] = 0
                    users_dict[user]['result_salary'] = 0
                    mess += 'Не работал;'
            except Exception as ex:
                logger.error(ex)
            users_dict[user]['comment'] = mess

            try:
                row = PaySheetModel.objects.get(user=user, week=week, year=year)
                users_dict[user]['bonus'] = row.bonus
                users_dict[user]['penalty'] = row.penalty
            except:
                users_dict[user]['bonus'] = 0
                users_dict[user]['penalty'] = 0

        sorted_dict = dict(sorted(users_dict.items(), key=lambda x: int(x[1]['hours']), reverse=True))
        context['users_dict'] = sorted_dict

        timedelta_tuples = (Appointment.objects.filter(date__year=year, date__week=week, verified=True)
                            .filter(filter_role2)
                            .values_list('duration'))
        timedelta_list = [item[0] for item in timedelta_tuples]
        total_hours = int(sum([i.total_seconds() for i in timedelta_list]) // 3600)
        total_result_salary = sum(item['result_salary'] for item in sorted_dict.values())

        context['total_hours'] = total_hours
        context['total_salary'] = total_salary
        context['total_result_salary'] = round(total_result_salary, 2)

        context['year'], context['week'] = year, week
        context['monday'], context['sunday'] = monday, sunday

        context['pay_sheets'] = PaySheetModel.objects.filter(year=year, week=week).filter(filter_role2)
        flag_button = True
        if current_week <= week and current_year == year:
            flag_button = False

        logger.success(flag_button)
        context['flag_button'] = flag_button
        logger.success(datetime.datetime.now() - time_start)
        pprint(context)
        return context


class PaySheetMonth(PaySheet):
    template_name = 'pay_sheet/pay_sheet_month.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        filter_role = Q(role__type_salary=Role.TYPE_SALARY[2][0])
        filter_role2 = Q(user__role__type_salary=Role.TYPE_SALARY[2][0])

        if not self.request.user.is_staff:
            raise Http404
        time_start = datetime.datetime.now()
        users_dict = {}
        total_salary = 0
        total_result_salary = 0
        today = datetime.date.today()
        current_year = today.year
        current_month = today.month

        year = self.request.GET.get('year', None)
        month = self.request.GET.get('month', None)
        if not year or not month:
            year = today.year
            month = today.month
        else:
            year, month = int(year), int(month)
        first_day, last_day, days = get_days_for_current_and_next_month(year, month)

        user_list = CustomUser.objects.filter(filter_role)

        for user in user_list:
            queryset_appointments = Appointment.objects.filter(user=user, date__year=year, date__month=month,
                                                               verified=True)
            if len(queryset_appointments) == 0:
                continue
            week_hours_work = []
            for day in days:
                try:
                    duration_day = sum([int(i.duration.total_seconds() / 3600) for i in
                                        queryset_appointments.filter(user=user, date=day)])
                    week_hours_work.append(duration_day)
                except Exception as ex:
                    logger.error(f'Нет записи на {day}')

            users_dict[user] = {'week_hours_work': week_hours_work}
            users_dict[user]['hours'] = sum(users_dict[user]['week_hours_work'])
            count_of_12 = users_dict[user]['week_hours_work'].count(12)
            users_dict[user]['count_of_12'] = count_of_12

            users_dict[user]['flag'] = True if count_of_12 >= 3 else False

            # Высчитывание зарплаты по часам
            salary = 0
            mess = ''

            if user.role.type_salary2 == 'Почасовая':
                for day_hour in users_dict[user]['week_hours_work']:
                    salary += day_hour * user.role.salary
            else:
                salary = user.role.salary

            users_dict[user]['salary'] = salary
            total_salary += salary
            # высчитывание коэффецента к зарплате
            try:
                users_dict[user]['works'] = {}
                users_dict[user]['kf'] = 0
                if user.role.calc_kf:
                    work_totals_dict = calculate_work_totals_for_week(user, users_dict[user]['hours'],
                                                                      first_day, last_day)
                    if work_totals_dict:
                        users_dict[user]['works'] = work_totals_dict
                        kf = sum([i[1] for i in work_totals_dict.values()]) * 100
                        users_dict[user]['kf'] = round(kf, 2)

                        if users_dict[user]['hours'] < 20:
                            mess += 'Отработанно меньше 20 часов; '
                        if kf >= 80:
                            result_salary = salary
                        else:
                            result_salary = salary * kf / 100
                            mess += 'Коэффецент меньше 80%; '

                        users_dict[user]['result_salary'] = round(result_salary, 2)
                    else:
                        users_dict[user]['result_salary'] = 0
                        mess = 'Нет сдельных листов; '

                else:
                    if user.role.type_salary2 == 'Почасовая':
                        users_dict[user]['result_salary'] = user.role.salary * users_dict[user]['hours']
                    else:
                        users_dict[user]['result_salary'] = user.role.salary
                if user.status_work == False:
                    mess += 'Уволен;'
            except Exception as ex:
                logger.error(ex)

            try:
                row = PaySheetMonthModel.objects.get(user=user, month=month, year=year)
                users_dict[user]['bonus'] = row.bonus
                users_dict[user]['penalty'] = row.penalty
                if row.kf > 110:
                    mess += 'Премия 5000 р.;'
            except:
                users_dict[user]['bonus'] = 0
                users_dict[user]['penalty'] = 0
                if users_dict[user]['kf'] > 110 and user.role.calc_kf and more_than_30_days_ago(user.date_joined.date(),
                                                                                                last_day):
                    users_dict[user]['bonus'] += 5000
                    mess += 'Премия 5000 р.;'
            users_dict[user]['comment'] = mess
            users_dict[user]['result_salary'] += users_dict[user]['bonus']
            users_dict[user]['result_salary'] -= users_dict[user]['penalty']
        sorted_dict = dict(sorted(users_dict.items(), key=lambda x: int(x[1]['hours']), reverse=True))
        context['users_dict'] = sorted_dict
        total_result_salary = sum(item['result_salary'] for item in sorted_dict.values())

        timedelta_tuples = (Appointment.objects.filter(date__month=month, verified=True)
                            .filter(filter_role2)
                            .values_list('duration'))
        timedelta_list = [item[0] for item in timedelta_tuples]
        total_hours = int(sum([i.total_seconds() for i in timedelta_list]) // 3600)
        context['total_hours'] = total_hours
        context['total_salary'] = total_salary
        context['total_result_salary'] = round(total_result_salary, 2)

        context['year'], context['month'] = year, month
        context['first_day'], context['last_day'] = first_day, last_day

        context['pay_sheets'] = PaySheetMonthModel.objects.filter(year=year, month=month)
        flag_button = True
        if current_month <= month and current_year == year:
            flag_button = False
            logger.debug('asdasdasd')
        context['flag_button'] = flag_button
        logger.success(datetime.datetime.now() - time_start)
        return context


def more_than_30_days_ago(date1, date2):
    # Парсинг строковых представлений дат в объекты datetime
    # date1 = datetime.datetime.strptime(date1, "%Y-%m-%d")
    # date2 = datetime.datetime.strptime(date2, "%Y-%m-%d")
    date_difference = abs(date1 - date2)
    if date_difference.days > 30:
        return True
    else:
        return False


def calculate_work_totals_for_week(user, hours, first_day, last_day):
    # Получите все записи работы за заданный период
    work_records = WorkRecord.objects.filter(user=user, delivery=None,
                                             date__range=(first_day, last_day), is_checked=True)

    # Используйте агрегацию, чтобы вычислить сумму работы для каждого вида работы
    work_totals = WorkRecordQuantity.objects.filter(work_record__in=work_records).values('standard').annotate(
        total_quantity=Sum('quantity'))

    # Создайте словарь, в котором ключами будут объекты модели Standards, а значениями - суммарное количество работы
    work_totals_dict = {}
    if work_totals:
        for work_total in work_totals:
            standard_id = work_total['standard']
            total_quantity = work_total['total_quantity']
            if total_quantity > 0:
                try:
                    standard = Standards.objects.get(pk=standard_id)
                    work_totals_dict[standard] = (total_quantity, total_quantity / (hours * standard.standard))
                except:
                    work_totals_dict['удаленные виды работ'] = (total_quantity, 0)
    return work_totals_dict


@login_required
@require_POST
def created_salary_check(request):
    try:
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
        data_str = list(request.POST.keys())[0]
        data_dict = json.loads(data_str)

        rowData = data_dict.get('rowData', [])
        year = data_dict.get('year', None)
        week = data_dict.get('week', None)
        month = data_dict.get('month', None)
        clean_data = map(clean_data_post, rowData)
        dict_pays = {}
        for row in clean_data:
            try:
                user = CustomUser.objects.get(username=row[0].split('тел:')[0])
                temp = {
                    'year': year,
                    'week': week,
                    'month': month,
                    'role': row[1],
                    'role_salary': row[2],
                    'hours': round(float(row[3]), 2) if row[3] != '' else 0,
                    'salary': round(float(row[4]), 2) if row[4] != '' else 0,
                    'works': row[5],
                    'count_of_12': int(row[6]) if row[6] != '' else 0,
                    'kf': round(float(row[7]), 2) if row[7] != '' else 0,
                    'result_salary': round(float(row[8]), 2) if row[8] != '' else 0,
                    'bonus': round(abs(float(row[9])), 2) if row[9] != '' else 0,
                    'penalty': round(abs(float(row[10])), 2) if row[10] != '' else 0,
                    'comment': row[11],
                }
                # temp['result_salary'] = round(temp['result_salary'] + temp['bonus'] - temp['penalty'], 2)
                dict_pays[user] = temp
            except Exception as ex:
                logger.error(ex)

        try:
            # Цикл для создания и сохранения объектов PaySheetModel
            for user, data in dict_pays.items():
                try:
                    # Проверяем существование записи для данного пользователя, недели и года
                    if not month:
                        pay_sheet = PaySheetModel.objects.get(user=user, year=data['year'], week=data['week'])
                    else:
                        pay_sheet = PaySheetMonthModel.objects.get(user=user, year=data['year'], month=data['month'])

                    # Обновляем поля записи данными из запроса
                    pay_sheet.role = data['role']
                    pay_sheet.role_salary = data['role_salary']
                    pay_sheet.hours = data['hours']
                    pay_sheet.salary = data['salary']
                    pay_sheet.works = data['works']
                    pay_sheet.count_of_12 = data['count_of_12']
                    pay_sheet.kf = data['kf']
                    pay_sheet.result_salary = data['result_salary']
                    pay_sheet.bonus = data['bonus']
                    pay_sheet.penalty = data['penalty']
                    pay_sheet.comment = data['comment']
                    pay_sheet.save()
                except Exception as ex:
                    logger.error(ex)
                    # Если запись не существует, создаем новую
                    if not month:
                        pay_sheet = PaySheetModel(
                            user=user,
                            year=data['year'],
                            week=data['week'],
                            role=data['role'],
                            role_salary=data['role_salary'],
                            hours=data['hours'],
                            salary=data['salary'],
                            works=data['works'],
                            count_of_12=data['count_of_12'],
                            kf=data['kf'],
                            result_salary=data['result_salary'],
                            bonus=data['bonus'],
                            penalty=data['penalty'],
                            comment=data['comment'],
                        )
                    else:
                        pay_sheet = PaySheetMonthModel(
                            user=user,
                            year=data['year'],
                            month=data['month'],
                            role=data['role'],
                            role_salary=data['role_salary'],
                            hours=data['hours'],
                            salary=data['salary'],
                            works=data['works'],
                            count_of_12=data['count_of_12'],
                            kf=data['kf'],
                            result_salary=data['result_salary'],
                            bonus=data['bonus'],
                            penalty=data['penalty'],
                            comment=data['comment'],
                        )
                    pay_sheet.save()

            return JsonResponse({'message': 'Успешно'})
        except Exception as ex:
            logger.error(ex)
            return JsonResponse({'message': 'Произошла ошибка'})

    except Exception as ex:
        logger.error(ex)
        return JsonResponse({'message': 'Произошла ошибка'})


def clean_data_post(row):
    output_row = [i.replace('\n', '')
                  .replace('  ', '')
                  .replace('Все работы', '')
                  .replace('р.', '')
                  .replace('%', '')
                  .replace('ч.', '')
                  .replace(',', '.')
                  .replace('руб/час', '').strip()
                  for i in row]
    return output_row


class PaySheetListView(LoginRequiredMixin, ListView):
    model = PaySheetModel
    template_name = 'pay_sheet/list_user_pay_sheets.html'
    context_object_name = 'pay_sheets'

    def get_queryset(self):
        current_user = self.request.user
        if current_user.role.type_salary == 'Раз в неделю':
            queryset = PaySheetModel.objects.filter(user=current_user).order_by('-year', '-week')
        else:
            queryset = PaySheetMonthModel.objects.filter(user=current_user).order_by('-year', '-month')
        return queryset


class PaySheetDetailView(LoginRequiredMixin, TemplateView):
    template_name = 'pay_sheet/pay_sheet_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = context['pk']
        model = context['model']
        if not self.test_func(pk, model):
            raise PermissionDenied
        if context['model'] == 'paysheetmonthmodel':
            pay_sheet = PaySheetMonthModel.objects.get(pk=pk)
        else:
            pay_sheet = PaySheetModel.objects.get(pk=pk)
            context['monday'], context['sunday'] = get_dates(pay_sheet.year, pay_sheet.week)
        context['pay_sheet'] = pay_sheet

        return context

    def test_func(self, pk, model):
        # Получите объект записи, к которой осуществляется доступ
        if model == 'paysheetmonthmodel':
            pay_sheet = PaySheetMonthModel.objects.get(pk=pk)
        else:
            pay_sheet = PaySheetModel.objects.get(pk=pk)

        # Проверьте, что текущий пользователь - владелец записи или имеет права is_staff
        return self.request.user == pay_sheet.user or self.request.user.is_staff
