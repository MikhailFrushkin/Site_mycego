import calendar
import json
import locale
from datetime import date, timedelta, datetime, time

from django.apps import apps
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import FormView, ListView, TemplateView
from loguru import logger
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from users.models import CustomUser, Role
from utils.utils import get_year_week
from work_schedule.forms import AppointmentForm, VacationRequestForm
from work_schedule.models import Appointment, VacationRequest, FingerPrint, BadFingerPrint
from django.core.exceptions import ValidationError


def format_duration(duration):
    hours, remainder = divmod(duration.seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}"


def ajax_view(request):
    if request.method == 'POST':
        selected_date = request.POST.get('date', None)

        if selected_date == 'current_month':
            today = date.today()
            # Вычисление начальной даты текущего месяца
            start_date = date(today.year, today.month, 1)
            # Вычисление конечной даты текущего месяца
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1)
            else:
                end_date = date(today.year, today.month + 1, 1)
            appointments = Appointment.objects.filter(date__gte=start_date, date__lt=end_date)
        elif selected_date == 'all':
            today = date.today()
            start_of_week = today - timedelta(days=today.weekday())
            appointments = Appointment.objects.filter(date__gte=start_of_week)
        elif selected_date == 'my':
            today = date.today()
            start_of_week = today - timedelta(days=today.weekday())
            appointments = Appointment.objects.filter(user_id=request.user.id, date__gte=start_of_week)
        else:
            appointments = Appointment.objects.filter(date=selected_date)

        if appointments:
            # Преобразуйте записи в список словарей
            appointments_list = [{"user": appointment.user.username, "date": appointment.date.strftime("%Y-%m-%d"),
                                  "start_time": appointment.start_time.strftime("%H:%M"),
                                  "end_time": appointment.end_time.strftime("%H:%M"),
                                  "duration": format_duration(appointment.duration),
                                  "verified": appointment.verified,
                                  "id": appointment.id,
                                  } for appointment in appointments]
        else:
            appointments_list = [{"user": 'пусто', "date": 'пусто',
                                  "start_time": 'пусто',
                                  "end_time": 'пусто',
                                  "duration": 'пусто',
                                  "verified": False,
                                  "id": 'пусто',
                                  }]
        response_data = {'appointments': appointments_list, 'user': request.user.username}
        return JsonResponse(response_data)


@login_required  # Ensure the user is logged in
@require_POST  # Accept only POST requests for this view
def delete_appointment(request):
    appointment_id = request.POST.get('id', None)
    appointment = get_object_or_404(Appointment, id=appointment_id)
    if appointment.user == request.user:
        appointment.delete()  # Delete the appointment
        return JsonResponse({'message': 'Appointment deleted successfully.'})
    else:
        return JsonResponse({'message': 'You do not have permission to delete this appointment.'}, status=403)


@login_required
@require_POST
def update_appointment(request):
    try:
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
        data_str = list(request.POST.keys())[0]
        data_dict = json.loads(data_str)

        type_m = data_dict.get('type', None)
        logger.debug(type_m)

        date_str = data_dict.get('date')
        month_names = {
            'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4,
            'мая': 5, 'июня': 6, 'июля': 7, 'августа': 8,
            'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
        }

        for month_name, month_number in month_names.items():
            if month_name in date_str:
                formatted_date_str = date_str.replace(month_name, str(month_number))
                date_obj = datetime.strptime(formatted_date_str, "%d %m %Y г.")
                break

        rowData = data_dict.get('rowData', [])
        appointment_list = []
        for data in rowData:
            unique_elements = {}

            for index, item in enumerate(data):
                if item != 'Нет' and item != 'Очистить':
                    if item in unique_elements:
                        unique_elements[item].append(index)
                    else:
                        unique_elements[item] = [index]
            if unique_elements:
                for key, value in unique_elements.items():
                    value.sort()
                    grouped_values = []
                    current_group = [value[0]]
                    for i in range(1, len(value)):
                        if value[i] == value[i - 1] + 1:
                            current_group.append(value[i])
                        else:
                            grouped_values.append(current_group)
                            current_group = [value[i]]
                    grouped_values.append(current_group)

                    for group in grouped_values:
                        start_time = 9 + group[0]
                        if len(group) == 1:
                            end_time = 10 + group[0]
                        else:
                            end_time = 10 + group[-1]

                        start_time = time(start_time, 0)
                        end_time = time(end_time, 0)
                        user = CustomUser.objects.get(username=key)

                        try:
                            comment = Appointment.objects.get(user=user, date=date_obj).comment
                        except:
                            comment = f"Утверждено в {datetime.now()}"
                        appointment = Appointment(
                            user=user,
                            date=date_obj,
                            start_time=start_time,
                            end_time=end_time,
                            verified=True,
                            comment=comment
                        )
                        appointment_list.append(appointment)
        if type_m:
            logger.debug(Appointment.objects.filter(date=date_obj, user__role__type_salary='Раз в месяц'))
            Appointment.objects.filter(date=date_obj, user__role__type_salary='Раз в месяц').delete()
        else:
            logger.debug(Appointment.objects.filter(date=date_obj, user__role__type_salary='Раз в неделю'))
            Appointment.objects.filter(date=date_obj, user__role__type_salary='Раз в неделю').delete()
        for item in appointment_list:
            item.save()
        return JsonResponse({'message': 'Appointment update successfully.'})

    except json.JSONDecodeError as e:
        return JsonResponse({'error': 'Invalid JSON data.'}, status=400)


def search_emp(request):
    try:
        logger.debug(request.GET)
        search = request.GET.get('searchTerm', '')
        week = request.GET.get('week', '')
        type_m = request.GET.get('type', '')
        logger.debug(search)
        logger.debug(week)
        logger.debug(type_m)

        # Создаем Q-объект для поиска по нику пользователя
        nickname_query = Q(user__username__icontains=search)

        # Создаем Q-объект для поиска по неделе
        week_query = Q(date__week=week)
        if type_m:
            role_query = Q(user__role__type_salary='Раз в месяц')
        else:
            role_query = Q(user__role__type_salary='Раз в неделю')
        # Комбинируем оба Q-объекта через оператор И (AND)
        query = Appointment.objects.filter(nickname_query & week_query & role_query).order_by('user', 'date')
        logger.debug(query)
        results = [{
            'username': appointment.user.username,
            'date': appointment.date,
            'start_time': appointment.start_time,
            'end_time': appointment.end_time,
            'duration': appointment.duration,
            'verified': appointment.verified,

        } for appointment in query]

        return JsonResponse({'results': results})
    except Exception as e:
        logger.error(e)
        return JsonResponse({'error': str(e)}, status=400)


def get_days_for_current_and_next_month():
    current_date = date.today()

    # Получаем первый день текущего месяца
    first_day_of_current_month = date(current_date.year, current_date.month, 1)

    # Получаем первый день следующего месяца
    if current_date.month == 12:
        first_day_of_next_month = date(current_date.year + 1, 1, 1)
    else:
        first_day_of_next_month = date(current_date.year, current_date.month + 1, 1)

    # Получаем последний день текущего месяца
    last_day_of_current_month = date(current_date.year, current_date.month,
                                     calendar.monthrange(current_date.year, current_date.month)[1])

    # Получаем последний день следующего месяца
    last_day_of_next_month = date(first_day_of_next_month.year, first_day_of_next_month.month,
                                  calendar.monthrange(first_day_of_next_month.year, first_day_of_next_month.month)[1])

    # Создаем список дней для текущего и следующего месяца
    days_for_current_and_next_month = []
    current_day = first_day_of_current_month
    while current_day <= last_day_of_next_month:
        days_for_current_and_next_month.append(current_day)
        current_day += timedelta(days=1)

    return days_for_current_and_next_month


class WorkSchedule(LoginRequiredMixin, FormView):
    template_name = 'work/work.html'
    login_url = '/users/login/'
    success_url = reverse_lazy('work:work_page')
    form_class = AppointmentForm
    context_object_name = 'appointments'
    paginate_by = 50  # Установите количество записей на странице

    def get(self, request, *args, **kwargs):
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        appointments = Appointment.objects.filter(date__gte=start_of_week)

        paginator = Paginator(appointments, self.paginate_by)
        page_number = request.GET.get('page')

        page = paginator.get_page(page_number)
        return render(
            request,
            self.template_name,
            {
                'form': self.get_form(),
                'appointments': page.object_list,
                'page': page,
                'user_role': self.request.user.role.type_salary
            }
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial']['user_id'] = self.request.user.id
        return kwargs

    def form_valid(self, form):
        current_user = self.request.user
        # Присваиваем пользователя к форме перед сохранением
        form.instance.user_id = current_user.id

        # Получаем данные из формы, чтобы передать их в функцию clean
        cleaned_data = form.cleaned_data

        # Ваша текущая логика из метода clean
        start_time = cleaned_data.get('start_time')
        end_time = cleaned_data.get('end_time')
        date = cleaned_data.get('date')
        user_id = cleaned_data.get('user_id')

        errors = []
        print('Проверяемая дата: ', cleaned_data)

        if date is None:
            errors.append(ValidationError("Не выбрана дата."))
        else:
            current_date = date.today()
            current_day_of_week = current_date.strftime('%A')
            current_week = current_date.isocalendar()[1]
            if current_day_of_week == 'Sunday':
                target_week = current_week + 1
            else:
                target_week = current_week

            start_date = current_date + timedelta(weeks=target_week - current_week, days=-current_date.weekday())
            days_in_target_week = [start_date + timedelta(days=i) for i in range(14)]
            end_day = start_date + timedelta(days=14)

            if self.request.user.role.type_salary == 'Раз в месяц':
                days_in_target_week = get_days_for_current_and_next_month()
                start_date = days_in_target_week[0]
                end_day = days_in_target_week[-1]

            if date not in days_in_target_week:
                errors.append(ValidationError(f"Вы можете записаться на даты, начиная с {start_date} до  {end_day}."))

        if start_time >= end_time:
            errors.append(ValidationError("Время начала должно быть меньше времени окончания."))

        existing_appointment = Appointment.objects.filter(
            Q(user_id=user_id),
            Q(date=date)
        )
        if existing_appointment.exists():
            errors.append(ValidationError("У вас уже есть запись на эту дату."))

        if errors:
            for error in errors:
                form.add_error(None, error)
            return render(self.request, self.template_name, {'form': form})

        form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def form_invalid(self, form):
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        appointments = Appointment.objects.filter(date__gte=start_of_week)
        return render(
            self.request,
            self.template_name,
            {
                'form': form,
                'appointments': appointments,  # Оставьте appointments в контексте
            }
        )


class GrafUser(LoginRequiredMixin, ListView):
    model = Appointment
    template_name = 'work/graf_user.html'
    login_url = '/users/login/'
    success_url = reverse_lazy('work:graf_user')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        selected_month = self.request.GET.get('selected_month')

        if selected_month and selected_month.isdigit():
            selected_month = int(selected_month)
        else:
            selected_month = today.month

        # Создаем объект datetime для первого дня указанного месяца
        start_of_month = datetime(today.year, selected_month, 1).date()

        # Находим последний день текущего месяца
        if selected_month == 12:
            end_of_month = datetime(today.year + 1, 1, 1).date()
        else:
            end_of_month = datetime(today.year, selected_month + 1, 1).date()

        # Вычитаем 1 день из последнего дня, чтобы получить последний день текущего месяца
        end_of_month -= timedelta(days=1)

        appointments = Appointment.objects.filter(
            Q(user=self.request.user),
            Q(date__gte=start_of_month),
            Q(date__lte=end_of_month)
        ).order_by('date')

        context['appointments'] = appointments
        return context


class EditWork(LoginRequiredMixin, TemplateView):
    template_name = 'work/edit_work.html'
    login_url = '/users/login/'
    success_url = reverse_lazy('work:edit_work')

    def create_excel_file(self, queryset, week):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Записи на работу'

        # Получите модель 'Appointment' из вашего приложения
        appointment_model = apps.get_model(app_label='work_schedule', model_name='Appointment')

        # Получите метаданные модели, чтобы получить поля
        model_fields = appointment_model._meta.fields

        # Добавьте столбец "Role" к заголовкам столбцов
        headers = [field.verbose_name for field in model_fields]
        headers.append('Должность')

        # Создайте заголовки столбцов на основе полей модели
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Заполните данные из queryset
        for row_num, appointment in enumerate(queryset, 2):
            for col_num, field in enumerate(model_fields, 1):
                cell_value = getattr(appointment, field.name)

                # Получите должность пользователя и добавьте ее в конец строки
                if field.name == 'user' and isinstance(cell_value, CustomUser):
                    worksheet.cell(row=row_num, column=col_num, value=cell_value.username)
                    worksheet.cell(row=row_num, column=len(headers), value=cell_value.role.name)
                else:
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)

                max_length = len(str(cell_value))
                column_letter = get_column_letter(col_num)
                if worksheet.column_dimensions[column_letter].width is None or worksheet.column_dimensions[
                    column_letter].width < max_length:
                    worksheet.column_dimensions[column_letter].width = max_length

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={week} week.xlsx'
        workbook.save(response)

        return response

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if 'download_excel' in request.GET:
            try:
                year = self.request.GET.get('year', None)
                week = self.request.GET.get('week', None)
                if not year or not week:
                    import datetime
                    today = datetime.date.today()
                    year = today.year
                    week = today.isocalendar()[1]

                queryset = Appointment.objects.filter(
                    date__year=year,
                    date__week=week
                )

            except:
                pass
            excel_response = self.create_excel_file(queryset, week)
            return excel_response

        context = self.get_context_data()
        return self.render_to_response(context)

    def get_queryset(self):
        if hasattr(self, '_queryset'):
            return self._queryset

        year = self.request.GET.get('year')
        week = self.request.GET.get('week')
        if not year or not week:
            import datetime
            today = datetime.date.today()
            year = today.year
            week = today.isocalendar()[1]
        if self.template_name == 'work/edit_work.html':
            queryset = Appointment.objects.filter(
                date__year=year,
                date__week=week,
                user__role__type_salary=Role.TYPE_SALARY[1][0]
            )
        elif self.template_name == 'work/edit_work_month.html':
            queryset = Appointment.objects.filter(
                date__year=year,
                date__week=week,
                user__role__type_salary=Role.TYPE_SALARY[2][0]
            )
        self._queryset = queryset  # Сохраняем результат в атрибуте _queryset
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        time_start = datetime.now()
        queryset = self.get_queryset().select_related('user__role')

        unique_dates = queryset.values_list('date', flat=True).distinct()
        unique_users = set(queryset.values_list('user', flat=True))
        not_role = ['Директор',
                    'Директор по производству'
                    ]
        if self.template_name == 'work/edit_work.html':
            none_write_graf = (CustomUser.objects.filter(status_work=True, role__type_salary='Раз в неделю')
                               .exclude(pk__in=unique_users)
                               .exclude(role__name__in=not_role).order_by('role', 'username'))
        else:
            none_write_graf = (CustomUser.objects.filter(status_work=True, role__type_salary='Раз в месяц')
                               .exclude(pk__in=unique_users)
                               .exclude(role__name__in=not_role).order_by('role', 'username'))
        context['none_write_graf'] = none_write_graf

        work_schedule = {}
        for index, date_appointment in enumerate(unique_dates):
            user_dict = {}
            appointments = queryset.filter(date=date_appointment)

            role_dict = {}
            total_hours = 0
            flag = True

            for appointment in appointments:
                role = appointment.user.role
                hours_role = int(format_duration(appointment.duration).split(':')[0])

                if role:
                    if role.name not in role_dict:
                        role_dict[role.name] = [1, hours_role]
                    else:
                        role_dict[role.name][0] += 1
                        role_dict[role.name][1] += hours_role
                    total_hours += hours_role

                if not appointment.verified:
                    flag = False

                work_hours = [0] * 12
                start_hour = appointment.start_time.hour
                end_hour = appointment.end_time.hour
                for i in range(start_hour - 9, end_hour - 9):
                    work_hours[i] = 1

                user_key = (
                    appointment.user.username, appointment.verified, appointment.user.role, appointment.user.avg_kf)
                if user_key in user_dict:
                    existing_work_hours = user_dict[user_key]
                    updated_work_hours = [a | b for a, b in zip(existing_work_hours, work_hours)]
                    user_dict[user_key] = updated_work_hours
                else:
                    user_dict[user_key] = work_hours
            sorted_user_dict = dict(sorted(user_dict.items(), key=lambda item: item[0][2].name))
            work_hours_count = {hour: sum(user_work_hours[hour] for user_work_hours in user_dict.values()) for hour in
                                range(12)}

            work_schedule[(date_appointment, f'table-{index}')] = (
                sorted_user_dict, work_hours_count, flag, role_dict, total_hours
            )

        context['work_schedule'] = work_schedule

        if self.template_name == 'work/edit_work.html':
            context['users'] = (CustomUser.objects.filter(status_work=True, role__type_salary='Раз в неделю')
                                .exclude(role__name__in=not_role).distinct().order_by('username'))
        else:
            context['users'] = (CustomUser.objects.filter(status_work=True, role__type_salary='Раз в месяц')
                                .exclude(role__name__in=not_role).distinct().order_by('username'))

        context['users_add'] = json.dumps([user.username for user in context['users']])
        context['year'], context['week'] = get_year_week(self.request.GET, type='list_work')
        logger.success(datetime.now() - time_start)
        return context


class EditWorkMonth(EditWork):
    template_name = 'work/edit_work_month.html'


class VacationRequestCreateView(LoginRequiredMixin, View):
    template_name = 'work/vacation_request_form.html'
    form_class = VacationRequestForm

    def get(self, request):
        form = self.form_class()
        user = request.user
        vacations = VacationRequest.objects.filter(employee=user)
        return render(request, self.template_name, {'form': form, 'vacations': vacations})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            vacation_request = form.save(commit=False)
            vacation_request.employee = request.user
            overlapping_requests = VacationRequest.objects.filter(
                employee=vacation_request.employee,
                start_date__lte=request.POST['end_date'],
                end_date__gte=request.POST['start_date']
            )
            logger.debug(overlapping_requests)
            if len(overlapping_requests) > 0:
                messages.error(self.request, "В выбранные даты уже существуют записи на отпуск.")

            else:
                vacation_request.save()
                return redirect('work:create_vacation_request')
        vacations = VacationRequest.objects.filter(employee=request.user)
        return render(request, self.template_name, {'form': form, 'vacations': vacations})


class VacationRequestAdmin(LoginRequiredMixin, TemplateView):
    template_name = 'work/vacation_request_admin.html'
    form_class = VacationRequestForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vacations = VacationRequest.objects.all()
        result = {}
        month_dict = {
            1: "Январь",
            2: "Февраль",
            3: "Март",
            4: "Апрель",
            5: "Май",
            6: "Июнь",
            7: "Июль",
            8: "Август",
            9: "Сентябрь",
            10: "Октябрь",
            11: "Ноябрь",
            12: "Декабрь"
        }
        for vacation in vacations:
            # Извлекаем месяц и год из начальной даты
            month = vacation.start_date.month
            year = vacation.start_date.year

            # Получаем роль сотрудника
            role = vacation.employee.role

            # Создаем ключ для словаря result
            key = (month_dict[month], year)

            # Если ключ уже существует, добавляем продолжительность отпуска
            if key in result:
                if role in result[key]["roles"]:
                    result[key]["roles"][role] += vacation.duration
                else:
                    result[key]["roles"][role] = vacation.duration

                if role in result[key]["vacations"]:
                    result[key]["vacations"][role].append(vacation)
                else:
                    result[key]["vacations"][role] = [vacation]

            else:
                result[key] = {
                    "roles": {role: vacation.duration},
                    "vacations": {role: [vacation]}
                }

        sorted_result = dict(sorted(result.items(), key=lambda item: (item[0][1], item[0][0]), reverse=True))

        context['vacation_data'] = sorted_result

        return context


def delete_vacation(request, vacation_id):
    vacation = get_object_or_404(VacationRequest, pk=vacation_id)

    if request.method == 'DELETE':
        vacation.delete()
        return JsonResponse({'message': 'Заявка успешно удалена'})

    return redirect('work:vacation_admin')


def confirm_vacation(request, vacation_id):
    vacation = get_object_or_404(VacationRequest, pk=vacation_id)

    if request.method == 'POST':
        vacation.is_checked = True
        vacation.save()
        return JsonResponse({'message': 'Заявка успешно подтверждена'})

    return redirect('work:vacation_admin')


class FingerPrintView(LoginRequiredMixin, TemplateView):
    template_name = 'work/finger_print.html'
    login_url = '/users/login/'
    success_url = reverse_lazy('work:finger_print')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        year, week = get_year_week(self.request.GET, type='finger')
        context['year'], context['week'] = year, week
        users = CustomUser.objects.filter(status_work=True)
        current_datetime = timezone.now()
        # date_condition = Q(date__lt=current_datetime.date())
        fingerprint_dates = FingerPrint.objects.filter(date__week=week).values_list('date', flat=True).distinct()
        logger.debug(fingerprint_dates)
        output_dict = {}
        for user in users:
            # appointments = Appointment.objects.filter(user=user, date__week=week).filter(date_condition,
            #                                                                              date__in=fingerprint_dates)
            # fingerprints = FingerPrint.objects.filter(user=user, date__week=week).filter(date_condition,
            #                                                                              date__in=fingerprint_dates)
            appointments = Appointment.objects.filter(user=user, date__week=week).filter(date__in=fingerprint_dates)
            fingerprints = FingerPrint.objects.filter(user=user, date__week=week).filter(date__in=fingerprint_dates)
            for appointment in appointments:
                date = appointment.date
                if date not in output_dict:
                    output_dict[date] = {}
                if user not in output_dict[date]:
                    output_dict[date][user] = {'graf': [], 'scan': []}
                output_dict[date][user]['graf'].append(appointment.start_time)
                output_dict[date][user]['graf'].append(appointment.end_time)

            for fingerprint in fingerprints:
                date = fingerprint.date
                if date not in output_dict:
                    output_dict[date] = {}
                if user not in output_dict[date]:
                    output_dict[date][user] = {'graf': [], 'scan': []}
                output_dict[date][user]['scan'].append(fingerprint.time)

        for date, user_data in output_dict.items():
            for user, data in user_data.items():
                data['error']: list = []
                if data['graf']:
                    data['graf'] = [min(data['graf']), max(data['graf'])]
                else:
                    data['error'].append('Нет в графике')

                if data['scan']:
                    if len(data['scan']) == 1:
                        data['error'].append('1 отметка за день')
                    else:
                        data['scan'] = [min(data['scan']), max(data['scan'])]

                        if data['graf']:
                            if data['scan'][0] > data['graf'][0]:
                                time1 = data['graf'][0]
                                time2 = data['scan'][0]
                                datetime1 = datetime(2000, 1, 1, time1.hour, time1.minute, time1.second)
                                datetime2 = datetime(2000, 1, 1, time2.hour, time2.minute, time2.second)
                                time_difference = datetime2 - datetime1
                                time_difference_as_time = int(time_difference.total_seconds() // 60)
                                if time_difference_as_time > 15:
                                    data['error'].append(f"Опоздал на {time_difference_as_time} минут")

                            if data['scan'][1] < data['graf'][1]:
                                time1 = data['graf'][1]
                                time2 = data['scan'][1]
                                datetime1 = datetime(2000, 1, 1, time1.hour, time1.minute, time1.second)
                                datetime2 = datetime(2000, 1, 1, time2.hour, time2.minute, time2.second)
                                time_difference = datetime1 - datetime2
                                time_difference_as_time = int(time_difference.total_seconds() // 60)
                                if time_difference_as_time > 15:
                                    data['error'].append(f"Ушел раньше на {time_difference_as_time} минут")
                else:
                    data['error'].append('Нет отметок на сканере')
                if data['error']:
                    temp, _ = BadFingerPrint.objects.get_or_create(
                        user=user,
                        date=date,
                        comment=','.join(data['error'])
                    )
            output_dict[date] = dict(sorted(user_data.items(), key=lambda item: bool(item[1]['error']), reverse=True))
        sorted_dict = dict(sorted(output_dict.items(), key=lambda x: x[0]))

        context['output_dict'] = sorted_dict
        return context
